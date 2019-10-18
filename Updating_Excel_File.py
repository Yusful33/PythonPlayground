import openpyxl

theFile = openpyxl.load_workbook('90Days.xlsx')
allSheetNames = theFile.sheetnames
print(theFile.sheetnames)
currentSheet = theFile['Sheet1']
#print(currentSheet[input(What)].value)

#This loops through the first eight columns and prints out the value in those cells
for row in range(1, currentSheet.max_row + 1):
        #print(row)
        for column in "ABCDEFGH":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            #print(cell_name)
            print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
print(currentSheet[cell_name].value)


    
def find_specific_cell():
#    list = []
#    day = 0
    day = input("What day of the streak is today?")
#    list.append(day)
    for row in range(1, currentSheet.max_row + 1):
        for column in "ABCDEFGH":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == day:
                #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                return cell_name
            
find_specific_cell()

